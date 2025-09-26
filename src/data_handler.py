# src/data_handler.py
# -*- coding: utf-8 -*-
"""
Carga y parseo del Excel original:
- Instrucciones (sheet 'Instrucciones')
- Cuestionario (sheet 'Cuestionario'): preguntas A..J con opciones 3/2/1
- Niveles (sheets 'Nivel 1','Nivel 2','Nivel 3'): NIVEL, DEFINICION, CARACTERISTICAS, RUTA
- Recomendaciones (sheet 'Recomendaciones')
- Umbrales desde la fórmula en 'Cuestionario'!C81 si existe (fallback a 15/23)
"""

from __future__ import annotations
from pathlib import Path
from typing import Dict, Any, List, Optional
import re
import unicodedata

import pandas as pd
from openpyxl import load_workbook

# Nombre EXACTO del archivo esperado
EXCEL_FILENAME = "Cuestionario de autodiagnóstico en inclusión laboral LGBTIQ para agencias de empleo.xlsx"

DEFAULT_THRESHOLDS = {"nivel_1_max": 15, "nivel_2_max": 23}


def _norm_text(x: Any) -> str:
    return str(x).strip() if x is not None else ""


def _sheet_to_dataframe(path: Path, sheet: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet, dtype=str, engine="openpyxl")
    df = df.dropna(how="all").dropna(axis=1, how="all").fillna("")
    return df


def _looks_like_question(text: Any) -> bool:
    s = _norm_text(text)
    return ("¿" in s or "?" in s) and len(s) > 10


def _is_question_id(val: Any) -> bool:
    s = _norm_text(val)
    return len(s) == 1 and s.isalpha()


def _find_excel_path(data_dir: Path) -> Optional[Path]:
    # Intenta el nombre exacto primero
    exact = data_dir / EXCEL_FILENAME
    if exact.exists():
        return exact
    # Intento robusto ante normalización unicode (NFC/NFD) en macOS
    for p in data_dir.glob("*.xlsx"):
        if unicodedata.normalize("NFC", p.name) == unicodedata.normalize(
            "NFC", EXCEL_FILENAME
        ):
            return p
    # Como último recurso, usa el primer .xlsx que haya
    xs = list(data_dir.glob("*.xlsx"))
    return xs[0] if xs else None


# ---------------- Instrucciones ----------------


def _load_instructions_from_excel(xlsx: Path) -> str:
    df = _sheet_to_dataframe(xlsx, "Instrucciones")
    parts: List[str] = []
    for _, row in df.iterrows():
        for v in row.tolist():
            sv = _norm_text(v)
            if sv:
                parts.append(sv)
    # Quita repeticiones triviales preservando orden
    seen, uniq = set(), []
    for p in parts:
        if p not in seen:
            uniq.append(p)
            seen.add(p)
    return (
        "\n".join(uniq)
        if uniq
        else "Bienvenido. Por favor lea y complete el cuestionario."
    )


# ---------------- Cuestionario ----------------


def _load_questions_from_excel(xlsx: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(xlsx, data_only=False, read_only=True)
    if "Cuestionario" not in wb.sheetnames:
        raise FileNotFoundError("La hoja 'Cuestionario' no existe en el Excel.")
    ws = wb["Cuestionario"]

    def cell(col: str, r: int):
        return ws[f"{col}{r}"].value

    questions: List[Dict[str, Any]] = []
    current_section: Optional[str] = None
    row, max_row = 1, ws.max_row

    while row <= max_row:
        b, c, d = cell("B", row), cell("C", row), cell("D", row)

        # Encabezado de sección (texto en B, C vacío o no-pregunta)
        if (
            isinstance(b, str)
            and not _is_question_id(b)
            and (c in (None, "") or not _looks_like_question(c))
        ):
            sb = _norm_text(b)
            if sb and sb.lower() != "respuesta":
                current_section = sb

        # Fila de pregunta (ID en B, pregunta en C)
        if _is_question_id(b) and _looks_like_question(c):
            qid = _norm_text(b)
            qtext = _norm_text(c)
            # Lee 3 opciones siguientes: B ∈ {3,2,1}, C = texto
            options: List[Dict[str, Any]] = []
            r2 = row + 1
            while r2 <= max_row:
                b2, c2 = cell("B", r2), cell("C", r2)
                if isinstance(b2, (int, float)) or (
                    isinstance(b2, str) and b2.isdigit()
                ):
                    score = int(str(b2))
                    if score in (3, 2, 1) and _norm_text(c2):
                        options.append({"score": score, "label": _norm_text(c2)})
                        r2 += 1
                        continue
                break
            # Asegura orden 3-2-1 y completa si faltan
            options = sorted(options, key=lambda x: x["score"], reverse=True)
            if len(options) != 3:
                missing = {3, 2, 1} - {o["score"] for o in options}
                for s in sorted(missing, reverse=True):
                    options.append({"score": s, "label": f"Opción {s}"})
                options = sorted(options, key=lambda x: x["score"], reverse=True)

            questions.append(
                {
                    "id": qid,
                    "section": current_section or "",
                    "text": qtext,
                    "options": options,
                }
            )
            row = r2
            continue

        row += 1

    if not questions:
        raise ValueError(
            "No se encontraron preguntas en 'Cuestionario'. Verifica el formato."
        )
    return questions


# ---------------- Umbrales desde fórmula C81 ----------------


def _extract_thresholds_from_formula(xlsx: Path) -> Dict[str, int]:
    try:
        wb = load_workbook(xlsx, data_only=False, read_only=False)
        ws = wb["Cuestionario"]
        f = ws["C81"].value
        if not isinstance(f, str) or "IF(" not in f.upper():
            return DEFAULT_THRESHOLDS.copy()
        comps = re.findall(r"D77\s*([<>=]{1,2})\s*(\d+)", f)
        nums = sorted({int(n) for (_op, n) in comps if n.isdigit()})
        if len(nums) >= 2:
            return {"nivel_1_max": nums[0], "nivel_2_max": nums[1]}
        if len(nums) == 1:
            return {
                "nivel_1_max": nums[0],
                "nivel_2_max": DEFAULT_THRESHOLDS["nivel_2_max"],
            }
        return DEFAULT_THRESHOLDS.copy()
    except Exception:
        return DEFAULT_THRESHOLDS.copy()


# ---------------- Niveles ----------------


def _load_single_level_sheet(xlsx: Path, sheet_name: str) -> Dict[str, str]:
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb[sheet_name]
    out: Dict[str, str] = {}
    for row in ws.iter_rows():
        cells = [c for c in row if c.value not in (None, "")]
        if len(cells) < 2:
            continue
        left = str(cells[0].value).strip().upper()
        right = str(cells[1].value).strip()
        mapping = {
            "NIVEL": "NIVEL",
            "DEFINICION": "DEFINICION",
            "DEFINICIÓN": "DEFINICION",
            "CARACTERISTICAS": "CARACTERISTICAS",
            "CARACTERÍSTICAS": "CARACTERISTICAS",
            "RUTA": "RUTA",
            "RUTA DE APRENDIZAJE SUGERIDA": "RUTA",
            "RUTA_DE_APRENDIZAJE_SUGERIDA": "RUTA",
        }
        key = mapping.get(left)
        if key and right:
            out[key] = f"{out[key]}\n{right}" if key in out else right
    if not out:
        df = _sheet_to_dataframe(xlsx, sheet_name)
        for col in df.columns:
            cu = str(col).strip().upper()
            if cu in [
                "NIVEL",
                "DEFINICION",
                "DEFINICIÓN",
                "CARACTERISTICAS",
                "CARACTERÍSTICAS",
                "RUTA",
                "RUTA_DE_APRENDIZAJE_SUGERIDA",
            ]:
                text = "\n".join(
                    [_norm_text(x) for x in df[col].tolist() if _norm_text(x)]
                )
                cu = "DEFINICION" if cu == "DEFINICIÓN" else cu
                cu = "CARACTERISTICAS" if cu == "CARACTERÍSTICAS" else cu
                out[cu] = text
    return out


def _load_levels_from_excel(xlsx: Path) -> Dict[str, Dict[str, str]]:
    out = {}
    for sn in ["Nivel 1", "Nivel 2", "Nivel 3"]:
        try:
            out[sn] = _load_single_level_sheet(xlsx, sn)
        except Exception:
            out[sn] = {}
    return out


# ---------------- Recomendaciones ----------------


def _normalize_rec_columns(cols: List[str]) -> List[str]:
    normed = []
    for c in cols:
        s = _norm_text(c).lower()
        s = (
            s.replace("á", "a")
            .replace("é", "e")
            .replace("í", "i")
            .replace("ó", "o")
            .replace("ú", "u")
            .replace("ñ", "n")
        )
        s = s.replace(" ", "_")
        normed.append(s)
    return normed


def _load_recommendations_from_excel(xlsx: Path) -> pd.DataFrame:
    df = _sheet_to_dataframe(xlsx, "Recomendaciones")
    df.columns = _normalize_rec_columns(df.columns.tolist())
    mapping = {}
    for c in df.columns:
        if "barrera" in c:
            mapping[c] = "barrera"
        elif "concepto" in c:
            mapping[c] = "concepto"
        elif "sintoma" in c:
            mapping[c] = "sintomas"
        elif "indicador" in c:
            mapping[c] = "indicadores"
        elif "recomendacion" in c:
            mapping[c] = "recomendaciones"
    if mapping:
        df = df.rename(columns=mapping)
    return df


# ---------------- API pública ----------------


def load_data_from_excel(excel_path: str | Path) -> Dict[str, Any]:
    xlsx = Path(excel_path)
    if not xlsx.exists():
        raise FileNotFoundError(f"No encontré el Excel en: {xlsx}")
    instructions = _load_instructions_from_excel(xlsx)
    questions = _load_questions_from_excel(xlsx)
    thresholds = _extract_thresholds_from_formula(xlsx)
    levels = _load_levels_from_excel(xlsx)
    recs = _load_recommendations_from_excel(xlsx)
    return {
        "instructions": instructions,
        "questions": questions,
        "thresholds": thresholds,
        "levels": levels,
        "recommendations": recs,
    }


def load_data(data_dir: str | Path = "data") -> Dict[str, Any]:
    d = Path(data_dir)
    xlsx = _find_excel_path(d)
    if not xlsx:
        raise FileNotFoundError(
            f"No encontré el Excel '{EXCEL_FILENAME}' en {d}. "
            "Súbelo con el uploader o colócalo en la carpeta 'data/'."
        )
    return load_data_from_excel(xlsx)
